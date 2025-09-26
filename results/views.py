from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.http import HttpResponse, Http404
from django.core.exceptions import ValidationError
from datetime import datetime, timedelta
import json
import csv
import io
import os
import zipfile
from openpyxl import Workbook

from .models import CollectionResult, ResultAnalysis, DataExport
from .serializers import (
    CollectionResultSerializer, CollectionResultCreateSerializer, CollectionResultUpdateSerializer,
    ResultAnalysisSerializer, ResultAnalysisCreateSerializer, ResultAnalysisUpdateSerializer,
    DataExportSerializer, DataExportCreateSerializer, DataExportUpdateSerializer,
    ResultStatsSerializer, ResultBatchOperationSerializer, ResultExportSerializer,
    ResultSearchSerializer
)
from devices.models import Device
from tasks.models import CollectionTask
from core.models import OperationLog
from core.utils import get_client_ip, get_user_agent


class CollectionResultListView(generics.ListAPIView):
    """采集结果列表视图"""
    serializer_class = CollectionResultSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = CollectionResult.objects.select_related(
            'task', 'device', 'task_execution'
        ).order_by('-created_at')
        
        # 搜索过滤
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(task__name__icontains=search) |
                Q(device__name__icontains=search) |
                Q(device__ip_address__icontains=search)
            )
        
        # 任务过滤
        task_id = self.request.query_params.get('task_id')
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        
        # 设备过滤
        device_id = self.request.query_params.get('device_id')
        if device_id:
            queryset = queryset.filter(device_id=device_id)
        
        # 状态过滤
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # 数据类型过滤
        data_type = self.request.query_params.get('data_type')
        if data_type:
            queryset = queryset.filter(data_type=data_type)
        
        # 处理状态过滤
        processed = self.request.query_params.get('processed')
        if processed is not None:
            queryset = queryset.filter(processed=processed.lower() == 'true')
        
        # 归档状态过滤
        archived = self.request.query_params.get('archived')
        if archived is not None:
            queryset = queryset.filter(archived=archived.lower() == 'true')
        
        # 时间范围过滤
        start_time = self.request.query_params.get('start_time')
        end_time = self.request.query_params.get('end_time')
        if start_time:
            queryset = queryset.filter(created_at__gte=start_time)
        if end_time:
            queryset = queryset.filter(created_at__lte=end_time)
        
        return queryset


class CollectionResultDetailView(generics.RetrieveUpdateDestroyAPIView):
    """采集结果详情视图"""
    queryset = CollectionResult.objects.select_related('task', 'device', 'task_execution')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'PUT' or self.request.method == 'PATCH':
            return CollectionResultUpdateSerializer
        return CollectionResultSerializer
    
    def perform_update(self, serializer):
        serializer.save()
        
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            resource_type='collection_result',
            resource_id=self.get_object().id,
            resource_name=f"{self.get_object().task.name} - {self.get_object().device.name}",
            details=f"更新采集结果",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
    
    def perform_destroy(self, instance):
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='delete',
            resource_type='collection_result',
            resource_id=instance.id,
            resource_name=f"{instance.task.name} - {instance.device.name}",
            details=f"删除采集结果",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
        
        instance.delete()


class ResultAnalysisListCreateView(generics.ListCreateAPIView):
    """结果分析列表创建视图"""
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = ResultAnalysis.objects.select_related(
            'task', 'created_by'
        ).prefetch_related('devices').order_by('-created_at')
        
        # 搜索过滤
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(task__name__icontains=search)
            )
        
        # 分析类型过滤
        analysis_type = self.request.query_params.get('analysis_type')
        if analysis_type:
            queryset = queryset.filter(analysis_type=analysis_type)
        
        # 状态过滤
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # 任务过滤
        task_id = self.request.query_params.get('task_id')
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        
        return queryset
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return ResultAnalysisCreateSerializer
        return ResultAnalysisSerializer
    
    def perform_create(self, serializer):
        analysis = serializer.save(created_by=self.request.user)
        
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            resource_type='result_analysis',
            resource_id=analysis.id,
            resource_name=analysis.name,
            details=f"创建结果分析: {analysis.analysis_type}",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )


class ResultAnalysisDetailView(generics.RetrieveUpdateDestroyAPIView):
    """结果分析详情视图"""
    queryset = ResultAnalysis.objects.select_related('task', 'created_by').prefetch_related('devices')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'PUT' or self.request.method == 'PATCH':
            return ResultAnalysisUpdateSerializer
        return ResultAnalysisSerializer
    
    def perform_update(self, serializer):
        serializer.save()
        
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            resource_type='result_analysis',
            resource_id=self.get_object().id,
            resource_name=self.get_object().name,
            details=f"更新结果分析",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
    
    def perform_destroy(self, instance):
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='delete',
            resource_type='result_analysis',
            resource_id=instance.id,
            resource_name=instance.name,
            details=f"删除结果分析",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
        
        instance.delete()


class DataExportListCreateView(generics.ListCreateAPIView):
    """数据导出列表创建视图"""
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = DataExport.objects.select_related(
            'task', 'created_by'
        ).prefetch_related('devices').order_by('-created_at')
        
        # 搜索过滤
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(task__name__icontains=search)
            )
        
        # 导出格式过滤
        export_format = self.request.query_params.get('export_format')
        if export_format:
            queryset = queryset.filter(export_format=export_format)
        
        # 状态过滤
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # 任务过滤
        task_id = self.request.query_params.get('task_id')
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        
        return queryset
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DataExportCreateSerializer
        return DataExportSerializer
    
    def perform_create(self, serializer):
        export = serializer.save(created_by=self.request.user)
        
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            resource_type='data_export',
            resource_id=export.id,
            resource_name=export.name,
            details=f"创建数据导出: {export.export_format}",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )


class DataExportDetailView(generics.RetrieveUpdateDestroyAPIView):
    """数据导出详情视图"""
    queryset = DataExport.objects.select_related('task', 'created_by').prefetch_related('devices')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'PUT' or self.request.method == 'PATCH':
            return DataExportUpdateSerializer
        return DataExportSerializer
    
    def perform_update(self, serializer):
        serializer.save()
        
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            resource_type='data_export',
            resource_id=self.get_object().id,
            resource_name=self.get_object().name,
            details=f"更新数据导出",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
    
    def perform_destroy(self, instance):
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            action='delete',
            resource_type='data_export',
            resource_id=instance.id,
            resource_name=instance.name,
            details=f"删除数据导出",
            ip_address=get_client_ip(self.request),
            user_agent=get_user_agent(self.request)
        )
        
        # 删除导出文件
        if instance.file_path and os.path.exists(instance.file_path):
            os.remove(instance.file_path)
        
        instance.delete()


class DataExportDownloadView(APIView):
    """数据导出下载视图"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        try:
            export = DataExport.objects.get(pk=pk)
        except DataExport.DoesNotExist:
            raise Http404("导出记录不存在")
        
        # 检查文件是否存在
        if not export.file_path or not os.path.exists(export.file_path):
            return Response(
                {"error": "导出文件不存在或已过期"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 检查是否过期
        if export.is_expired:
            return Response(
                {"error": "导出文件已过期"},
                status=status.HTTP_410_GONE
            )
        
        # 更新下载统计
        export.download_count += 1
        export.last_download_at = timezone.now()
        export.save(update_fields=['download_count', 'last_download_at'])
        
        # 返回文件
        with open(export.file_path, 'rb') as f:
            response = HttpResponse(f.read())
            
        # 设置响应头
        filename = os.path.basename(export.file_path)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 根据文件格式设置Content-Type
        if export.export_format == 'csv':
            response['Content-Type'] = 'text/csv'
        elif export.export_format == 'json':
            response['Content-Type'] = 'application/json'
        elif export.export_format == 'excel':
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            response['Content-Type'] = 'application/octet-stream'
        
        # 记录操作日志
        OperationLog.objects.create(
            user=request.user,
            action='download',
            resource_type='data_export',
            resource_id=export.id,
            resource_name=export.name,
            details=f"下载导出文件",
            ip_address=get_client_ip(request),
            user_agent=get_user_agent(request)
        )
        
        return response


class ResultBatchOperationView(APIView):
    """结果批量操作视图"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = ResultBatchOperationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        operation = serializer.validated_data['operation']
        result_ids = serializer.validated_data['result_ids']
        options = serializer.validated_data.get('options', {})
        
        results = CollectionResult.objects.filter(id__in=result_ids)
        success_count = 0
        error_count = 0
        errors = []
        
        for result in results:
            try:
                if operation == 'process':
                    result.processed = True
                    result.processed_at = timezone.now()
                    result.processor = request.user.username
                    result.save(update_fields=['processed', 'processed_at', 'processor'])
                    success_count += 1
                
                elif operation == 'archive':
                    result.archived = True
                    result.archive_path = options.get('archive_path', '')
                    result.save(update_fields=['archived', 'archive_path'])
                    success_count += 1
                
                elif operation == 'delete':
                    result.delete()
                    success_count += 1
                
                elif operation == 'export':
                    # 这里可以添加导出逻辑
                    success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"结果 {result.id}: {str(e)}")
        
        # 记录操作日志
        OperationLog.objects.create(
            user=request.user,
            action='batch_operation',
            resource_type='collection_result',
            resource_id=None,
            resource_name=f"批量{operation}",
            details=f"批量{operation} {len(result_ids)}个结果，成功{success_count}个，失败{error_count}个",
            ip_address=get_client_ip(request),
            user_agent=get_user_agent(request)
        )
        
        return Response({
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def result_stats(request):
    """结果统计API"""
    # 基础统计
    total_results = CollectionResult.objects.count()
    success_results = CollectionResult.objects.filter(status='success').count()
    failed_results = CollectionResult.objects.filter(status='failed').count()
    partial_results = CollectionResult.objects.filter(status='partial').count()
    timeout_results = CollectionResult.objects.filter(status='timeout').count()
    error_results = CollectionResult.objects.filter(status='error').count()
    
    # 数据类型统计
    text_results = CollectionResult.objects.filter(data_type='text').count()
    json_results = CollectionResult.objects.filter(data_type='json').count()
    xml_results = CollectionResult.objects.filter(data_type='xml').count()
    csv_results = CollectionResult.objects.filter(data_type='csv').count()
    binary_results = CollectionResult.objects.filter(data_type='binary').count()
    
    # 处理状态统计
    processed_results = CollectionResult.objects.filter(processed=True).count()
    unprocessed_results = CollectionResult.objects.filter(processed=False).count()
    archived_results = CollectionResult.objects.filter(archived=True).count()
    
    # 分析统计
    total_analysis = ResultAnalysis.objects.count()
    pending_analysis = ResultAnalysis.objects.filter(status='pending').count()
    running_analysis = ResultAnalysis.objects.filter(status='running').count()
    completed_analysis = ResultAnalysis.objects.filter(status='completed').count()
    failed_analysis = ResultAnalysis.objects.filter(status='failed').count()
    
    # 导出统计
    total_exports = DataExport.objects.count()
    pending_exports = DataExport.objects.filter(status='pending').count()
    processing_exports = DataExport.objects.filter(status='processing').count()
    completed_exports = DataExport.objects.filter(status='completed').count()
    failed_exports = DataExport.objects.filter(status='failed').count()
    
    # 7天趋势数据
    daily_trends = []
    for i in range(7):
        date = timezone.now().date() - timedelta(days=i)
        daily_results = CollectionResult.objects.filter(
            created_at__date=date
        ).aggregate(
            total=Count('id'),
            success=Count('id', filter=Q(status='success')),
            failed=Count('id', filter=Q(status='failed'))
        )
        
        daily_trends.append({
            'date': date.strftime('%Y-%m-%d'),
            'total': daily_results['total'] or 0,
            'success': daily_results['success'] or 0,
            'failed': daily_results['failed'] or 0,
            'success_rate': round(
                (daily_results['success'] or 0) / max(daily_results['total'] or 1, 1) * 100, 2
            )
        })
    
    data = {
        'total_results': total_results,
        'success_results': success_results,
        'failed_results': failed_results,
        'partial_results': partial_results,
        'timeout_results': timeout_results,
        'error_results': error_results,
        'text_results': text_results,
        'json_results': json_results,
        'xml_results': xml_results,
        'csv_results': csv_results,
        'binary_results': binary_results,
        'processed_results': processed_results,
        'unprocessed_results': unprocessed_results,
        'archived_results': archived_results,
        'total_analysis': total_analysis,
        'pending_analysis': pending_analysis,
        'running_analysis': running_analysis,
        'completed_analysis': completed_analysis,
        'failed_analysis': failed_analysis,
        'total_exports': total_exports,
        'pending_exports': pending_exports,
        'processing_exports': processing_exports,
        'completed_exports': completed_exports,
        'failed_exports': failed_exports,
        'daily_trends': daily_trends
    }
    
    serializer = ResultStatsSerializer(data)
    return Response(serializer.data)


class ResultExportView(APIView):
    """结果导出视图"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = ResultExportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # 获取查询参数
        export_format = serializer.validated_data['format']
        task_id = serializer.validated_data.get('task_id')
        device_id = serializer.validated_data.get('device_id')
        status_filter = serializer.validated_data.get('status')
        data_type = serializer.validated_data.get('data_type')
        start_time = serializer.validated_data.get('start_time')
        end_time = serializer.validated_data.get('end_time')
        include_raw_data = serializer.validated_data.get('include_raw_data', False)
        include_parsed_data = serializer.validated_data.get('include_parsed_data', True)
        
        # 构建查询集
        queryset = CollectionResult.objects.select_related('task', 'device')
        
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        if device_id:
            queryset = queryset.filter(device_id=device_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if data_type:
            queryset = queryset.filter(data_type=data_type)
        if start_time:
            queryset = queryset.filter(created_at__gte=start_time)
        if end_time:
            queryset = queryset.filter(created_at__lte=end_time)
        
        queryset = queryset.order_by('-created_at')
        
        # 限制导出数量
        if queryset.count() > 10000:
            return Response(
                {"error": "导出数据量过大，请缩小查询范围"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 生成导出文件
        if export_format == 'csv':
            response = self._export_csv(queryset, include_raw_data, include_parsed_data)
        elif export_format == 'json':
            response = self._export_json(queryset, include_raw_data, include_parsed_data)
        elif export_format == 'excel':
            response = self._export_excel(queryset, include_raw_data, include_parsed_data)
        else:
            return Response(
                {"error": "不支持的导出格式"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 记录操作日志
        OperationLog.objects.create(
            user=request.user,
            action='export',
            resource_type='collection_result',
            resource_id=None,
            resource_name=f"结果导出({export_format})",
            details=f"导出{queryset.count()}条结果数据",
            ip_address=get_client_ip(request),
            user_agent=get_user_agent(request)
        )
        
        return response
    
    def _export_csv(self, queryset, include_raw_data, include_parsed_data):
        """导出CSV格式"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [
            'ID', '任务名称', '设备名称', '设备IP', '状态', '数据类型',
            '数据大小', '记录数', '错误数', '警告数', '已处理', '已归档', '创建时间'
        ]
        if include_raw_data:
            headers.append('原始数据')
        if include_parsed_data:
            headers.append('解析数据')
        
        writer.writerow(headers)
        
        # 写入数据
        for result in queryset:
            row = [
                result.id,
                result.task.name,
                result.device.name,
                result.device.ip_address,
                result.get_status_display(),
                result.get_data_type_display(),
                result.data_size or 0,
                result.record_count or 0,
                result.error_count or 0,
                result.warning_count or 0,
                '是' if result.processed else '否',
                '是' if result.archived else '否',
                result.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            if include_raw_data:
                row.append(result.raw_data or '')
            if include_parsed_data:
                row.append(result.parsed_data or '')
            
            writer.writerow(row)
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="collection_results.csv"'
        return response
    
    def _export_json(self, queryset, include_raw_data, include_parsed_data):
        """导出JSON格式"""
        data = []
        for result in queryset:
            item = {
                'id': result.id,
                'task_name': result.task.name,
                'device_name': result.device.name,
                'device_ip': result.device.ip_address,
                'status': result.status,
                'data_type': result.data_type,
                'data_size': result.data_size,
                'record_count': result.record_count,
                'error_count': result.error_count,
                'warning_count': result.warning_count,
                'processed': result.processed,
                'archived': result.archived,
                'created_at': result.created_at.isoformat()
            }
            
            if include_raw_data:
                item['raw_data'] = result.raw_data
            if include_parsed_data:
                item['parsed_data'] = result.parsed_data
            
            data.append(item)
        
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="collection_results.json"'
        return response
    
    def _export_excel(self, queryset, include_raw_data, include_parsed_data):
        """导出Excel格式"""
        wb = Workbook()
        ws = wb.active
        ws.title = "采集结果"
        
        # 写入表头
        headers = [
            'ID', '任务名称', '设备名称', '设备IP', '状态', '数据类型',
            '数据大小', '记录数', '错误数', '警告数', '已处理', '已归档', '创建时间'
        ]
        if include_raw_data:
            headers.append('原始数据')
        if include_parsed_data:
            headers.append('解析数据')
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row, result in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=result.id)
            ws.cell(row=row, column=2, value=result.task.name)
            ws.cell(row=row, column=3, value=result.device.name)
            ws.cell(row=row, column=4, value=result.device.ip_address)
            ws.cell(row=row, column=5, value=result.get_status_display())
            ws.cell(row=row, column=6, value=result.get_data_type_display())
            ws.cell(row=row, column=7, value=result.data_size or 0)
            ws.cell(row=row, column=8, value=result.record_count or 0)
            ws.cell(row=row, column=9, value=result.error_count or 0)
            ws.cell(row=row, column=10, value=result.warning_count or 0)
            ws.cell(row=row, column=11, value='是' if result.processed else '否')
            ws.cell(row=row, column=12, value='是' if result.archived else '否')
            ws.cell(row=row, column=13, value=result.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            col = 14
            if include_raw_data:
                ws.cell(row=row, column=col, value=result.raw_data or '')
                col += 1
            if include_parsed_data:
                ws.cell(row=row, column=col, value=result.parsed_data or '')
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="collection_results.xlsx"'
        return response
